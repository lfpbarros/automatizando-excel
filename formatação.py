from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook

# Padrão das classes:
font = Font(
    name='Calibri',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None, # alinhamento vertical
    underline='none',
    strike=False, # riscado
    color='FF000000'
)

fill = PatternFill(
    fill_type=None,
    start_color='FFFFFFFF',
    end_color='FF000000'
)

border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style='double',
                            color='FF0000FF'),
                            # o padrão do right é semelhante aos outros, alterei aqui só para facilitar a vida
                 top=Side(border_style=None,
                          color='FF000000'),
                 bottom=Side(border_style=None,
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )
# No border_style, usar um entre: 
# {‘thin’, ‘dashed’, ‘mediumDashDot’, ‘dashDotDot’, ‘hair’, 
# ‘dotted’, ‘mediumDashDotDot’, ‘medium’, ‘double’, ‘dashDot’, 
# ‘slantDashDot’, ‘thick’, ‘mediumDashed’}

alignment=Alignment(
    horizontal='general',
    vertical='bottom',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0)

number_format = 'General'

protection = Protection(
    locked=True,
    hidden=False)


wb = Workbook()

sheet = wb.active

# Alterando a fonte:
sheet['A1'] = 'teste'
sheet['A2'] = 'texto'
sheet['A3'] = 'text'
sheet['A1'].font = Font(bold=True, size=20, underline='single')
sheet['A2'].font = Font(italic=True, color='FFFF0000') #RGB

# Alterando o fundo:
sheet['A2'].fill = PatternFill(fill_type = 'solid', start_color='FF000000', end_color='FFFFF00F')

# Alterando a borda:
sheet['A3'].border = border

wb.save(r'formatação.xlsx')  