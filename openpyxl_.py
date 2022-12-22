from openpyxl import load_workbook

wb = load_workbook('exemplo.xlsx')

wb.get_sheet_names()
wb.sheetnames
# pegando os nomes das planilhas

wb.get_sheet_by_name('Sheet1')
sheet = wb['Sheet1']
# selecionando uma planilha

sheet['A3'].value
# exibindo um valor

sheet['B2'].value

sheet.cell(row=2, column=2).value
# outra forma de selecionar apenas um valor

sheet.max_row
# retorna o máximo de linhas 
sheet.max_column
# retorna o máximo de colunas

for i in range(1, sheet.max_row + 1):
    print(sheet.cell(row=i, column=2).value)

for i in range(1, sheet.max_row + 1):
    print(sheet['B' + str(i)].value)


sheet.cell(row=2, column=3).value = 75
# alterando um valor, para que isso se exiba na planilha, precisamos salvar ela.

wb.save('exemplo.xlsx')

# =======================
# Agrupamento:
sheet.merge_cells('A1:D1')
# mesclando as células
sheet.unmerge_cells('A1:D1')

wb.save('exemplo.xlsx')

sheet.insert_rows(4) # insere uma linha na posição escolhida
sheet.delete_rows(4) # deleta a linha na posição escolhida
# essas funções também servem para colunas

sheet.delete_cols(2, 5) # começa a excluir na segunda e avança 5, nesse caso da B:F

wb.save('exemplo2.xlsx')

# Adição de imagem:
from openpyxl.drawing.image import Image

img = Image('catlogo.png') # carregar imagem

sheet.add_image(img, 'A1')
wb.save('exemplo2.xlsx')
